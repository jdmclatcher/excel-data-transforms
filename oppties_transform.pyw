# Applies Data Transformations to Excel Reports

# Jonathan McLatcher
# Summer 2021

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
import tkinter as tk
from tkinter.filedialog import askopenfilename
from tkinter import font as tkFont
import os

def main():
    # establish GUI
    root = tk.Tk()
    root.title("Opportunities Report Data Enhancer")
    
    # create font
    button_font = tkFont.Font(weight='bold')

    # set window size
    root.geometry("586x230")

    # create buttons
    stage_1_button = tk.Button(root, text="STAGE 1", height=5, width=25, fg="blue", font=button_font, command=None)
    stage_1_button['state'] = tk.DISABLED

    stage_2_button = tk.Button(root, text="STAGE 2", height=5, width=25, fg="blue", font=button_font, command=None)
    stage_2_button['state'] = tk.DISABLED

    # button that can perform a function at the beginning
    pick_btn = tk.Button(root, text="CHOOSE FILE", height=5, width=25, font=button_font, command=lambda: pick_file(stage_1_button, stage_2_button, response_label, root))

    # reports status in GUI
    response_label = tk.Label(root, height=6, width=100, text="Wating for '.xlsx' file...", wraplength=550, fg="black")

    # scale gui with a grid
    root.grid_columnconfigure(0, weight=1)
    root.grid_columnconfigure(1, weight=1)
    root.grid_columnconfigure(2, weight=1)
    root.grid_rowconfigure(0, weight=1)
    root.grid_rowconfigure(1, weight=1)

    stage_1_button.grid_configure(sticky="nsew")
    stage_2_button.grid_configure(sticky="nsew")
    pick_btn.grid_configure(sticky="nsew")
    response_label.grid_configure(sticky="nsew")

    # setup rows and cols
    pick_btn.grid(row=0, column=0, padx=5, pady=7)
    stage_1_button.grid(row=0, column=1, padx=5, pady=7)
    stage_2_button.grid(row=0, column=2, padx=5, pady=7)
    response_label.grid(row=1, columnspan=3)

    root.mainloop()


def pick_file(stage_1_button, stage_2_button, response_label, root):
    try:
        # open file
        file_path = askopenfilename(title="Select a spreadsheet...")
        dir_path = os.path.dirname(file_path)
        # check for valid extension (".xlsx")
        if os.path.splitext(os.path.basename(file_path))[1] == ".xlsx":
            stage_1_button['state'] = tk.NORMAL
            stage_1_button.configure(command=lambda: check_convention(file_path, dir_path, response_label, stage_1_button, stage_2_button))
            stage_2_button['state'] = tk.NORMAL
            stage_2_button.configure(command=lambda: check_list_and_add(file_path, dir_path, response_label, stage_1_button, stage_2_button))
            response_label.configure(text="File Selected: '" + os.path.basename(file_path) + "'", fg="black")
        else:
            response_label.configure(text="ERROR: Please provide a valid '.xlsx' file.", fg="red")
            stage_1_button['state'] = tk.DISABLED
            stage_2_button['state'] = tk.DISABLED
    # fix for crashes when browsing for file
    except:
        root.destroy()
    finally:
        root.mainloop()


# STAGE 1 - ensure that (throughout the entire opportunity name) a " - " only occurs once
def check_convention(file_path, dir_path, response_label, stage_1_button, stage_2_button):
    workbook = load_workbook(file_path, data_only=True)
    data_sheet = workbook.active

    OPP_NAME_IDX = 3

    err_rows = [] # stores all the 'problematic' rows that need manual fixing
    row_num = 1 # reference to current row num
    for row in data_sheet.iter_rows(values_only=True):
        # stop if reached the empty line immediately before the metadata
        if row[0] == None:
            break
        if row[OPP_NAME_IDX - 1] != None:
            # store the header/title row in the 'new sheet' (copy-paste of the header row)
            if row_num == 1:
                err_rows.append(row)
            # convert to a string
            elif str(row[OPP_NAME_IDX - 1]).count(" - ") != 1:
                err_rows.append(row) # add to invalid rows sheet
        # increase the row number count for tracking
        row_num += 1

    # store 'err_rows' to new sheet
    workbook.create_sheet("To Be Fixed Entries")
    to_be_fixed_sheet = workbook["To Be Fixed Entries"]
    for row in err_rows:
        # appends all 'err_rows' to  
        to_be_fixed_sheet.append(row)

    # save
    # finds the correct path/directory to save the new export file to 
    new_path = dir_path + "/" + os.path.splitext(os.path.basename(file_path))[0] + "_processed.xlsx"
    # check to see if file is open in excel
    try:
        workbook.save(filename=new_path)
        # update GUI response label
        response_label.configure(text="(Stage 1) File Exported to: '" + new_path + "'", fg="green")
    except:
        # update GUI response label
        response_label.configure(text="(Stage 1) Failed to save file. Is the file currently open in Excel?", fg="red")
        return
    finally:
        # disable the buttons
        stage_1_button['state'] = tk.DISABLED
        stage_2_button['state'] = tk.DISABLED

# STEP 2 - after having received the 'cleaned up' data (only 1 " - " per entry),
#   check the second half of the entry with the naming conventions list and confirm that
#   they are in references data
def check_list_and_add(file_path, dir_path, response_label, stage_1_button, stage_2_button):

    MANAGER = 'A'
    MANAGER_IDX = 1
    LEADER = 'B'
    ACCOUNT_NAME = 'C'
    OPP_NAME = 'D'
    OPP_NAME_IDX = 4
    PRODUCT_NAME = 'E'
    PRODUCT_CATEGORY = 'F'
    PRODUCT_PARENT = 'G'
    STAGE = 'H'
    STAGE_IDX = 8
    ALGO = 'I'
    AMOUNT_CURRENCY = 'J'
    AMOUNT = 'K'
    AMOUNT_IDX = 11
    WEIGHTED = 'L'
    QUARTER = 'M'
    MONTH = 'N'
    CLOSE_DATE = 'O'
    CLOSE_DATE_IDX = 15

    # open all ws
    references_path = os.path.join(dir_path, "oppties_reference_data.xlsx")
    references_wb = load_workbook(references_path, data_only=True)
    products_sheet = references_wb["Products"]
    algo_sheet = references_wb["Algo"]
    leader_sheet = references_wb["Leader"]

    data_wb = load_workbook(file_path, data_only=True)
    data_sheet = data_wb.active

    data_sheet[MANAGER + "1"] = "MANAGER"

    # add cols and headings
    # columns are inserted BEFORE the provided index
    data_sheet.insert_cols(idx=MANAGER_IDX+1, amount=1)
    data_sheet[LEADER + "1"] = "Leader"
    data_sheet[LEADER + "1"].font = Font(bold=True)
    data_sheet[LEADER + "1"].alignment = Alignment(vertical='center', horizontal="center")

    data_sheet.insert_cols(idx=OPP_NAME_IDX+1, amount=3)
    data_sheet[PRODUCT_NAME + "1"] = "Product Name"
    data_sheet[PRODUCT_NAME + "1"].font = Font(bold=True)
    data_sheet[PRODUCT_NAME + "1"].alignment = Alignment(vertical='center', horizontal="center")
    
    data_sheet[PRODUCT_CATEGORY + "1"] = "Product Category"
    data_sheet[PRODUCT_CATEGORY + "1"].font = Font(bold=True)
    data_sheet[PRODUCT_CATEGORY + "1"].alignment = Alignment(vertical='center', horizontal="center")

    data_sheet[PRODUCT_PARENT + "1"] = "Product Parent"
    data_sheet[PRODUCT_PARENT + "1"].font = Font(bold=True)
    data_sheet[PRODUCT_PARENT + "1"].alignment = Alignment(vertical='center', horizontal="center")

    data_sheet.insert_cols(idx=STAGE_IDX+1, amount=1)
    data_sheet[ALGO + "1"] = "Algo"
    data_sheet[ALGO + "1"].font = Font(bold=True)
    data_sheet[ALGO + "1"].alignment = Alignment(vertical='center', horizontal="center")
    
    data_sheet.insert_cols(idx=AMOUNT_IDX+1, amount=3)
    data_sheet[WEIGHTED + "1"] = "Weighted"
    data_sheet[WEIGHTED + "1"].font = Font(bold=True)
    data_sheet[WEIGHTED + "1"].alignment = Alignment(vertical='center', horizontal="center")

    data_sheet[QUARTER + "1"] = "Quarter"
    data_sheet[QUARTER + "1"].font = Font(bold=True)
    data_sheet[QUARTER + "1"].alignment = Alignment(vertical='center', horizontal="center")

    data_sheet[MONTH + "1"] = "Month"
    data_sheet[MONTH + "1"].font = Font(bold=True)
    data_sheet[MONTH + "1"].alignment = Alignment(vertical='center', horizontal="center")

    # change col widths
    data_sheet.column_dimensions[MANAGER].width = 17.8
    data_sheet.column_dimensions[LEADER].width = 13.8
    data_sheet.column_dimensions[ACCOUNT_NAME].width = 30
    data_sheet.column_dimensions[OPP_NAME].width = 0
    data_sheet.column_dimensions[PRODUCT_PARENT].width = 12.3
    data_sheet.column_dimensions[STAGE].width = 25.5
    data_sheet.column_dimensions[ALGO].width = 8
    data_sheet.column_dimensions[AMOUNT_CURRENCY].width = 0
    data_sheet.column_dimensions[QUARTER].width = 8.5
    data_sheet.column_dimensions[MONTH].width = 6.5
    data_sheet.column_dimensions[CLOSE_DATE].width = 9.5


    # store rows here to report in new sheet
    err_rows = [] # store problematic rows (rows that aren't in reference data)
    row_num = 1 # start at row 1
    for row in data_sheet.iter_rows(values_only=True):
        # stop if reached the empty line immediately before the metadata
        if row[OPP_NAME_IDX - 1] == None:
            break # break = stop the for loop (end)
        # save title row
        if row_num == 1:
            err_rows.append(['1','2','3','4','5','6','7','8'])
        else:
            row_str = str(row_num)
            found = False
            for line in products_sheet.iter_rows(min_row=2, values_only=True):
                # skip if doesn't exist or has no other half
                # ERROR check if Stage 1 wasn't completed properly
                if len(row[OPP_NAME_IDX - 1].split(" - ")) == 1:
                    response_label.configure(text="(Stage 2) ERROR: Did you perhaps click 'Stage 2' on a file that is not ready for Stage 2?\nPlease try again.", fg="red")
                    # disable the buttons
                    stage_1_button['state'] = tk.DISABLED
                    stage_2_button['state'] = tk.DISABLED
                    return # stop
                else:
                    # splitting a string creates an array (0-indexed), then accessing '[1]' on that array will
                        # return the SECOND HALF of the string
                    if row[OPP_NAME_IDX - 1].split(" - ")[1] == line[0]:
                        found = True
                        # add data to cols if a filled conversion is available
                        if line[1] != None or line[2] != None or line[3] != None:
                            # WRITING DATA - using data_sheet and setting a certain [row, col] to a value using
                                # "=" (equals sign)
                            # line is the array reference of 1 line of the products reference data
                            data_sheet[PRODUCT_NAME + row_str] = line[1]
                            data_sheet[PRODUCT_CATEGORY + row_str] = line[2]
                            data_sheet[PRODUCT_PARENT + row_str] = line[3]
                        break # stop as soon as found
            # if never found, flag row and add to conversions sheet
            if found == False:
                err_rows.append(row)
                products_sheet.append([row[OPP_NAME_IDX - 1].split(" - ")[1]])

            if row[STAGE_IDX - 1] == algo_sheet.cell(row=2, column=2).value:
                # WRITES the corresponding algo value from that row to ALGO in data_sheet
                data_sheet[ALGO + row_str] = algo_sheet.cell(row=2, column=1).value
            elif row[STAGE_IDX - 1] == algo_sheet.cell(row=3, column=2).value:
                data_sheet[ALGO + row_str] = algo_sheet.cell(row=3, column=1).value
            elif row[STAGE_IDX - 1] == algo_sheet.cell(row=4, column=2).value:
                data_sheet[ALGO + row_str] = algo_sheet.cell(row=4, column=1).value
            elif row[STAGE_IDX - 1] == algo_sheet.cell(row=5, column=2).value:
                data_sheet[ALGO + row_str] = algo_sheet.cell(row=5, column=1).value
            elif row[STAGE_IDX - 1] == algo_sheet.cell(row=6, column=2).value:
                data_sheet[ALGO + row_str] = algo_sheet.cell(row=6, column=1).value
            elif row[STAGE_IDX - 1] == algo_sheet.cell(row=7, column=2).value:
                data_sheet[ALGO + row_str] = algo_sheet.cell(row=7, column=1).value
        
            # calculate weighted amount
            if row[AMOUNT_IDX - 1] != None and row[STAGE_IDX - 1] != None:
                data_sheet[WEIGHTED + row_str] = round(row[AMOUNT_IDX - 1] * data_sheet[ALGO + row_str].value, 2)

            # output quarter and year (YYYY-FQX) as well as month (10-Oct, ...)
            if row[CLOSE_DATE_IDX - 1] != None:
                year = row[CLOSE_DATE_IDX - 1].year # gets the year
                month = f"{row[CLOSE_DATE_IDX - 1]:%m}" # gets date with leading zeros
                letter_month = f"{row[CLOSE_DATE_IDX - 1]:%b}" # 3-letter month
                quarter = ""
                if month == "10" or month == "11" or month == "12":
                    quarter = "FQ1"
                elif month == "01" or month == "02" or month == "03":
                    quarter = "FQ2"
                elif month == "04" or month == "05" or month == "06":
                    quarter = "FQ3"
                elif month == "07" or month == "08" or month == "09":
                    quarter = "FQ4"
                
                # string concatenation and WRITE
                data_sheet[QUARTER + row_str] = str(year) + "-" + quarter
                data_sheet[MONTH + row_str] = month + "-" + letter_month
            
            acct_format = u'_($* #,##0_);[Red]_($* (#,##0);_($* -_0_0_);_(@'
            # set format to accounting
            data_sheet[WEIGHTED + row_str].number_format = acct_format
            data_sheet[AMOUNT + row_str].number_format = acct_format

            for mananger_line in leader_sheet.iter_rows(values_only=True):
                # when we are using row, we use the numbered index (numbers)
                # when we are using data_sheet, we use the column letter ('A', 'B', etc)
                if row[MANAGER_IDX - 1] == None:
                    break 
                
                if row[MANAGER_IDX - 1] == mananger_line[0]:
                    data_sheet[LEADER + row_str] = mananger_line[1]
                    break

        row_num += 1

    # store to new sheet
    data_wb.create_sheet("Missing Product Reference")
    no_product_sheet = data_wb["Missing Product Reference"]
    for row in err_rows:
        no_product_sheet.append(row)

    # save
    new_path = dir_path + "/" + os.path.splitext(os.path.basename(file_path))[0] + "_processed.xlsx"
    # check to see if file is open in excel
    try:
        data_wb.save(filename=new_path)
        references_wb.save(references_path)
        response_label.configure(text="(Stage 2) File Exported to: '" + new_path + "'", fg="green")
    except:
        response_label.configure(text="(Stage 2) Failed to save file(s). Is a file currently open in Excel?", fg="red")
        return
    finally:
        # disable the buttons
        stage_1_button['state'] = tk.DISABLED
        stage_2_button['state'] = tk.DISABLED


# runs "main" function at startup
if __name__ == "__main__":
    main()