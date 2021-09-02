# Applies Data Transformations to Excel Reports

# Jonathan McLatcher
# Summer 2021

import os
import platform
import tkinter as tk
from tkinter.filedialog import askopenfilename
from tkinter import font as tkFont
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from datetime import date
import csv


def main():
    root = tk.Tk()
    root.title("Report Data Enhancer")
    # for Max OS X
    if platform.system() == "Darwin":
        root.geometry("583x200")

        start_btn = tk.Button(root, text="START", height=5, width=20, fg="blue", command=None)
        start_btn['state'] = tk.DISABLED

        pick_btn = tk.Button(root, text="CHOOSE FILE", height=5, width=20, command=lambda: pick_file(start_btn, response_label, root))

        response_label = tk.Label(root, height=6, width=60, text="Wating for .xlsx file...", wraplength=500, fg="black")
    # for Windows
    else:
        # create font
        button_font = tkFont.Font(weight='bold')

        root.geometry("586x225")

        start_btn = tk.Button(root, text="START", height=5, width=30, fg="blue", font=button_font, command=None)
        start_btn['state'] = tk.DISABLED

        pick_btn = tk.Button(root, text="CHOOSE FILE", height=5, width=30, font=button_font, command=lambda: pick_file(start_btn, response_label, root))

        response_label = tk.Label(root, height=5, width=75, text="Wating for .xlsx file...", wraplength=500, fg="black")

    # scale gui
    root.grid_columnconfigure(0, weight=1)
    root.grid_columnconfigure(1, weight=1)
    root.grid_rowconfigure(0, weight=1)
    root.grid_rowconfigure(1, weight=1)

    start_btn.grid_configure(sticky="nsew")
    pick_btn.grid_configure(sticky="nsew")
    response_label.grid_configure(sticky="nsew")

    # setup rows and cols
    pick_btn.grid(row=0, column=0, padx=5, pady=7)
    start_btn.grid(row=0, column=1, padx=5, pady=7)
    response_label.grid(row=1, columnspan=3)

    root.mainloop()


def pick_file(start_btn, response_label, root):
    try:
        file_path = askopenfilename(title="Select a spreadsheet...")
        dir_path = os.path.dirname(file_path)
        # check for valid extension
        if os.path.splitext(os.path.basename(file_path))[1] == ".xlsx":
            start_btn['state'] = tk.NORMAL
            start_btn.configure(command=lambda: transform_data(file_path, dir_path, response_label, start_btn))
            response_label.configure(text="File Selected: '" + os.path.basename(file_path) + "'", fg="black")
        else:
            response_label.configure(text="ERROR: Please provide a valid '.xlsx' file.", fg="red")
            start_btn['state'] = tk.DISABLED
    # fix for crashes when browsing for file
    except:
        root.destroy()
    finally:
        root.mainloop()


def transform_data(file_path, dir_path, response_label, start_btn):
    workbook = load_workbook(file_path, data_only=True)
    data_sheet = workbook.active

    # base 1 for and indices - indices are just based off col letter
    # indices are all after the addition of the new rows
    ID_NUMBER = 'A'
    ID_NUMBER_IDX = 1
    REPORTED_EES = 'E'
    REPORTED_EES_IDX = 5
    EMPLOYEE_RANGE = 'F'
    PRODUCT_NAME = 'K'
    PRODUCT_NAME_IDX = 11
    PRODUCT_CATEGORY = 'L'
    COUNTRY = 'M'
    MIN_EES = 'N'
    TIER_MIN = 'O'
    TIER_MIN_IDX = 15
    TIER_MAX = 'P'
    TIER_MAX_IDX = 16
    EES_IN_TIER = 'Q'
    EES_IN_TIER_IDX = 17
    TIER_STATUS = 'R'
    TIER_COUNT = 'S'
    MAX_TIER = 'T'
    RATE_NUM = 'V'
    RATE_NUM_IDX = 22
    PERCENT_DIF = 'W'
    PERCENT_DIF_IDX = 23
    BLENDED_RATE = 'X'
    BLENDED_RATE_IDX = 24
    ONE_RATE = 'Y'
    ONE_RATE_IDX = 25
    NEXT_INCREASE = 'AE'
    NEXT_INCREASE_IDX = 31
    DAYS_UNTIL_INC = 'AF'
    DAYS_UNTIL_INC_IDX = 32

    MAX_ROWS = data_sheet.max_row

    # col for ee range
    data_sheet.insert_cols(idx=REPORTED_EES_IDX+1, amount=1)
    data_sheet[EMPLOYEE_RANGE + "1"] = "Employee Range"
    data_sheet[EMPLOYEE_RANGE + "1"].font = Font(bold=True)
    data_sheet[EMPLOYEE_RANGE + "1"].alignment = Alignment(vertical='center', horizontal="center")

    # cols for conversions (2)
    data_sheet.insert_cols(idx=PRODUCT_NAME_IDX+1, amount=2)
    data_sheet[PRODUCT_CATEGORY + "1"] = "Product Category"
    data_sheet[PRODUCT_CATEGORY + "1"].font = Font(bold=True)
    data_sheet[PRODUCT_CATEGORY + "1"].alignment = Alignment(vertical='center', horizontal="center")
    data_sheet[COUNTRY + "1"] = "Country"
    data_sheet[COUNTRY + "1"].font = Font(bold=True)
    data_sheet[COUNTRY + "1"].alignment = Alignment(vertical='center', horizontal="center")

    # cols for tiers/EEs (4)
    # col for num EEs per tier
    data_sheet.insert_cols(idx=TIER_MAX_IDX+1, amount=4)
    data_sheet[EES_IN_TIER + "1"] = "# of EEs in Tier"
    data_sheet[EES_IN_TIER + "1"].font = Font(bold=True)
    data_sheet[EES_IN_TIER + "1"].alignment = Alignment(vertical='center', horizontal="center")

    # col for tiered status
    data_sheet[TIER_STATUS + "1"] = "Tiered (Y/N)"
    data_sheet[TIER_STATUS + "1"].font = Font(bold=True)
    data_sheet[TIER_STATUS + "1"].alignment = Alignment(vertical='center', horizontal="center")

    # col for tier count + level
    data_sheet[TIER_COUNT + "1"] = "Tier Count + Level"
    data_sheet[TIER_COUNT + "1"].font = Font(bold=True)
    data_sheet[TIER_COUNT + "1"].alignment = Alignment(vertical='center', horizontal="center")

    # col for max tier (Y/N)
    data_sheet[MAX_TIER + "1"] = "Max Tier (Y/N)"
    data_sheet[MAX_TIER + "1"].font = Font(bold=True)
    data_sheet[MAX_TIER + "1"].alignment = Alignment(vertical='center', horizontal="center")

    # col for percent diff
    data_sheet.insert_cols(idx=RATE_NUM_IDX+1, amount=1)
    data_sheet[PERCENT_DIF + "1"] = "% Diff from Previous Tier"
    data_sheet[PERCENT_DIF + "1"].font = Font(bold=True)
    data_sheet[PERCENT_DIF + "1"].alignment = Alignment(vertical='center', horizontal="center")

    # col for blended rate
    data_sheet.insert_cols(idx=PERCENT_DIF_IDX+1, amount=1)
    data_sheet[BLENDED_RATE + "1"] = "Blended Rate"
    data_sheet[BLENDED_RATE + "1"].font = Font(bold=True)
    data_sheet[BLENDED_RATE + "1"].alignment = Alignment(vertical='center', horizontal="center")

    # col for one rate
    data_sheet.insert_cols(idx=BLENDED_RATE_IDX+1, amount=1)
    data_sheet[ONE_RATE + "1"] = "One Rate"
    data_sheet[ONE_RATE + "1"].font = Font(bold=True)
    data_sheet[ONE_RATE + "1"].alignment = Alignment(vertical='center', horizontal="center")

    # days until renewal
    data_sheet.insert_cols(idx=NEXT_INCREASE_IDX+1, amount=1)
    data_sheet[DAYS_UNTIL_INC + "1"] = "Days Until Renewal"
    data_sheet[DAYS_UNTIL_INC + "1"].font = Font(bold=True)
    data_sheet[DAYS_UNTIL_INC + "1"].alignment = Alignment(vertical='center', horizontal="center")

    prev_row = None
    tier_count = 0
    product_count = 1
    row_num = 2  # tracks the actual row - includes EXP rows
    tiered = False
    blended_total = 0

    # create array of all vals in conversion (template) sheet
    products = []
    conversions_path = os.path.join(dir_path, "subs_conversions.csv")
    with open(conversions_path, newline='') as products_csv:
        reader = csv.reader(products_csv, delimiter='|', quotechar='|')
        for line in reader:
            products.append(line)

    # NOTE: SUBTRACT 1 from IDX constant for correct access
    for row in data_sheet.iter_rows(min_row=2, values_only=True):
        row_str = str(row_num)
        # stop before reading metadata
        if row[ID_NUMBER_IDX - 1] == None or row[ID_NUMBER_IDX - 1] == "":
            break
        # 0. match product names to their respective categories
        # loop through product names
        for product in products:
            # if match found...
            if product[0] == row[PRODUCT_NAME_IDX - 1]:
                # add data to new cols
                data_sheet[PRODUCT_CATEGORY + row_str] = product[1]
                data_sheet[COUNTRY + row_str] = product[2]
                # stop as soon as match is found
                break

        # 0.5 report correct employee range based on number of reported EEs
        if row[REPORTED_EES_IDX - 1] >= 0 and row[REPORTED_EES_IDX - 1] <= 1500:
            data_sheet[EMPLOYEE_RANGE + row_str] = "A.1-1500"
        elif row[REPORTED_EES_IDX - 1] > 1500 and row[REPORTED_EES_IDX - 1] <= 5000:
            data_sheet[EMPLOYEE_RANGE + row_str] = "B.1501-5000"
        elif row[REPORTED_EES_IDX - 1] > 5000 and row[REPORTED_EES_IDX - 1] <= 10000:
            data_sheet[EMPLOYEE_RANGE + row_str] = "C.5001-10000"
        elif row[REPORTED_EES_IDX - 1] > 10000 and row[REPORTED_EES_IDX - 1] <= 20000:
            data_sheet[EMPLOYEE_RANGE + row_str] = "D.10001-20000"
        elif row[REPORTED_EES_IDX - 1] > 20000:
            data_sheet[EMPLOYEE_RANGE + row_str] = "E.20001+"
        else:
            print("Unexpected Error. Row #: " + row_str)

        # 1. determine and report if product is tiered (Y/N)
        #   if product has more than 1 row and an empty tier max, then tiered
        # 2. determine the tier count and level
        # peek next row
        next_row_str = str(row_num + 1)
        next_product = data_sheet[PRODUCT_NAME + next_row_str]
        # check for singles (non-tiered)
        if prev_row == None:
            # only check next
            if row[PRODUCT_NAME_IDX - 1] != next_product.value:
                # if was previously tiered, set as max
                if tiered == True:
                    data_sheet[MAX_TIER + row_str] = 'Y'
                tiered = False
            else:
                tiered = True
        else:
            # check both prev and next
            if prev_row[PRODUCT_NAME_IDX - 1] != row[PRODUCT_NAME_IDX - 1] and row[PRODUCT_NAME_IDX - 1] != next_product.value:
                tiered = False
            else:
                tiered = True
        if prev_row != None and prev_row[ID_NUMBER_IDX - 1] == row[ID_NUMBER_IDX - 1]:
            # same customer
            if prev_row[PRODUCT_NAME_IDX - 1] == row[PRODUCT_NAME_IDX - 1]:
                # same product
                if tier_count == 1:
                    tiered = True
                tier_count += 1
            else:
                # different product, same customer
                if tiered == True:
                    # give previous row the max tier attribute
                    product_count += 1
                    tier_count = 1
        else:
            # new customer
            product_count = 1
            tier_count = 1
        # FIXME changed this if statement - see backup for original
        if tiered and prev_row != None and row[TIER_MIN_IDX - 1] == prev_row[TIER_MIN_IDX - 1] and tier_count != 1:
            data_sheet[TIER_COUNT + row_str] = 'ERR'  # report ERR in col
            tier_count -= 1
        elif tiered:
            # add correct tier status
            data_sheet[TIER_STATUS + row_str] = 'Y'
            data_sheet[TIER_COUNT + row_str] = str(product_count) + "." + str(tier_count)

            # 3. determine correct number of employees per tier
            # 'None' case for tier min is revisited after
            if row[TIER_MIN_IDX - 1] != None:
                # check if tier max is 'infinite'/None OR for valid subtraction
                if row[TIER_MAX_IDX - 1] == None or row[TIER_MAX_IDX - 1] - row[TIER_MIN_IDX - 1] >= row[REPORTED_EES_IDX - 1] or tier_count > 1:
                    if row[TIER_MAX_IDX - 1] == None or row[TIER_MAX_IDX - 1] - row[TIER_MIN_IDX - 1] > row[REPORTED_EES_IDX - 1] - row[TIER_MIN_IDX - 1] + 1:
                        if row[REPORTED_EES_IDX - 1] - row[TIER_MIN_IDX - 1] > 0:
                            if tier_count == 1:
                                data_sheet[EES_IN_TIER + row_str] = row[REPORTED_EES_IDX - 1] - row[TIER_MIN_IDX - 1]
                            else:
                                data_sheet[EES_IN_TIER + row_str] = row[REPORTED_EES_IDX - 1] - row[TIER_MIN_IDX - 1] + 1
                    else:
                        data_sheet[EES_IN_TIER + row_str] = row[TIER_MAX_IDX - 1] - row[TIER_MIN_IDX - 1] + 1
                else:
                    # check if first tier
                    if row[TIER_MIN_IDX - 1] == 0:
                        data_sheet[EES_IN_TIER + row_str] = row[TIER_MAX_IDX - 1] - row[TIER_MIN_IDX - 1]
                    else:
                        # edge case where first tier min is greater than the number of EEs
                        if tier_count == 1 and row[TIER_MAX_IDX - 1] >= row[REPORTED_EES_IDX - 1]:
                            data_sheet[EES_IN_TIER + row_str] = row[REPORTED_EES_IDX - 1]
                        else:
                            data_sheet[EES_IN_TIER + row_str] = row[TIER_MAX_IDX - 1] - row[TIER_MIN_IDX - 1] + 1

            # check if mins and maxes are all 0 or None (or if EXP)
            if ((row[TIER_MIN_IDX - 1] == 0 or row[TIER_MIN_IDX - 1] == None) and (row[TIER_MAX_IDX - 1] == 0 or row[TIER_MAX_IDX - 1] == None)):
                data_sheet[TIER_COUNT + row_str] = 'ERR'  # report ERR in col
                tier_count -= 1
        else:
            # report EEs number from max reported EEs
            data_sheet[EES_IN_TIER + row_str] = row[REPORTED_EES_IDX - 1]
            # add tier status
            data_sheet[TIER_STATUS + row_str] = 'N'
            data_sheet[TIER_COUNT + row_str] = None
            if row[RATE_NUM_IDX - 1] != None:
                data_sheet[ONE_RATE + row_str] = row[RATE_NUM_IDX - 1]

        # 4. calculate 'blended rate' for each employees product
        #   calculate from EE count per tier and tier price(s)
        #   skipping any rows with 'ERR'
        #
        # CALCULATION:
        #   SUM OF (num EEs in tier * rate for tier) + (num EEs in next tier * rate for that tier) * ...
        #       ALL DIVIDED BY (total reported num EEs)
        #

        # peek previous row
        prev_row_str = str(row_num - 1)
        prev_leading_char = ''
        # special case for first row
        if row_num == 2:
            if tier_count == 1:  # if tiered
                blended_total += data_sheet[EES_IN_TIER + row_str].value * row[RATE_NUM_IDX - 1]
                data_sheet[MAX_TIER + row_str] = 'N'
            else:
                # not tiered
                data_sheet[MAX_TIER + row_str] = 'N'
                if row[RATE_NUM_IDX - 1] != None:
                    data_sheet[ONE_RATE + row_str] = row[RATE_NUM_IDX - 1]

        # special case for last row
        if row_num == MAX_ROWS:
            # if tiered and has EEs record
            if tier_count != 0 and data_sheet[EES_IN_TIER + row_str].value != None:
                blended_total += data_sheet[EES_IN_TIER + row_str].value * row[RATE_NUM_IDX - 1]
                # report final values for blended rate
                for i in range(int(tier_count)):
                    prev_row_str_temp = str(row_num - i)
                    data_sheet[BLENDED_RATE + prev_row_str_temp] = round(blended_total / row[REPORTED_EES_IDX - 1], 2)
                    data_sheet[ONE_RATE + prev_row_str_temp] = round(blended_total / row[REPORTED_EES_IDX - 1], 2)
                    data_sheet[MAX_TIER + prev_row_str_temp] = 'N'

                    # % DIFF
                    # for each single row in the tier,
                    # compare it to the previous one and find the dif
                    # that the current row has
                    cur_row_str_temp = str(row_num - i + 1)
                    if data_sheet[RATE_NUM + cur_row_str_temp].value != None and data_sheet[RATE_NUM + prev_row_str_temp].value != None and data_sheet[RATE_NUM + prev_row_str_temp].value != 0:
                        cur_rate = data_sheet[RATE_NUM + cur_row_str_temp].value
                        prev_rate = data_sheet[RATE_NUM + prev_row_str_temp].value
                        # skip first tier
                        if i != 0:
                            if prev_rate > cur_rate:
                                # make negative
                                data_sheet[PERCENT_DIF + cur_row_str_temp] = (1 - round(cur_rate / prev_rate, 2)) * -1
                            else:
                                data_sheet[PERCENT_DIF + cur_row_str_temp] = 1 - round(cur_rate / prev_rate, 2)

                data_sheet[MAX_TIER + row_str] = 'Y'
            else:
                # not tiered
                data_sheet[MAX_TIER + row_str] = 'N'
                if row[RATE_NUM_IDX - 1] != None:
                    data_sheet[ONE_RATE + row_str] = row[RATE_NUM_IDX - 1]
            break

        # if not None for anything or 0 for key numbers (if rate or total EEs are 0 for example)
        # skip calculation if anything is missing, or if it's the first or last row
        if data_sheet[TIER_COUNT + row_str].value != 'ERR' and (data_sheet[EES_IN_TIER + row_str].value != None and data_sheet[EES_IN_TIER + row_str].value != 0) and row[REPORTED_EES_IDX - 1] != 0 and row_num != 2 and row_num != MAX_ROWS:
            # if different substrings ('products') - check for NoneType prev row as well
            if data_sheet[TIER_COUNT + prev_row_str].value != None:
                prev_first = data_sheet[TIER_COUNT + prev_row_str].value.split('.', 1)[0]
            else:
                prev_first = None
            # account for changing from tiered to non-tiered
            if data_sheet[TIER_COUNT + row_str].value == None:
                first = None  # set first to None to indicate change in tiered type
            else:
                first = data_sheet[TIER_COUNT + row_str].value.split('.', 1)[0]
            if prev_first == None or prev_first != first:
                # product changed
                # set previous row to be max (only if tiered - meaning that prev_first would be some number)
                if prev_first != None:
                    data_sheet[MAX_TIER + prev_row_str] = 'Y'

                # finish calculation and apply to previous rows
                # divide current total by total EEs
                # apply to all previous rows in same tier (length of highest tier count and up)
                if (data_sheet[TIER_COUNT + prev_row_str].value != None and data_sheet[TIER_COUNT + prev_row_str].value != 'ERR'):
                    prev_last = data_sheet[TIER_COUNT + prev_row_str].value.split('.', 1)[1]
                    for i in range(int(prev_last)):
                        prev_row_str_temp = str(row_num - 1 - i)
                        # use previous row's EEs (to correspond with recurring calculations)
                        data_sheet[BLENDED_RATE + prev_row_str_temp] = round(blended_total / data_sheet[REPORTED_EES + prev_row_str].value, 2)
                        data_sheet[ONE_RATE + prev_row_str_temp] = round(blended_total / data_sheet[REPORTED_EES + prev_row_str].value, 2)

                        # % DIFF
                        # for each single row in the tier,
                        # compare it to the previous one and find the dif
                        # that the current row has
                        cur_row_str_temp = str(row_num - i)
                        if data_sheet[RATE_NUM + cur_row_str_temp].value != None and data_sheet[RATE_NUM + prev_row_str_temp].value != None and data_sheet[RATE_NUM + prev_row_str_temp].value != 0:
                            cur_rate = data_sheet[RATE_NUM + cur_row_str_temp].value
                            prev_rate = data_sheet[RATE_NUM + prev_row_str_temp].value
                            # skip first tier
                            if i != 0:
                                if prev_rate > cur_rate:
                                    # make negative
                                    data_sheet[PERCENT_DIF + cur_row_str_temp] = (1 - round(cur_rate / prev_rate, 2)) * -1
                                else:
                                    data_sheet[PERCENT_DIF + cur_row_str_temp] = 1 - round(cur_rate / prev_rate, 2)

                # reset running total
                blended_total = 0

                # then add the current row (first in tiered) - and add 'N' to row
                if data_sheet[EES_IN_TIER + row_str].value != None:
                    blended_total += data_sheet[EES_IN_TIER + row_str].value * row[RATE_NUM_IDX - 1]
                if tiered == True:
                    data_sheet[MAX_TIER + row_str] = 'N'
            else:
                # product didn't change, keep addding to calc
                # add EEs in tier * rate to running total
                if data_sheet[EES_IN_TIER + row_str].value != None:
                    blended_total += data_sheet[EES_IN_TIER + row_str].value * row[RATE_NUM_IDX - 1]
                # set current row as not the max
                data_sheet[MAX_TIER + row_str] = 'N'
        else:
            # update vals for ERR, None, 0 valued rows
            data_sheet[MAX_TIER + row_str] = 'N'

        # FIXME fix max tier not working when following row is ERR

        if row[NEXT_INCREASE_IDX - 1] != None:
            # calculate and report the number of days until contract renews (next increase date - current day)
            data_sheet[DAYS_UNTIL_INC + row_str] = (row[NEXT_INCREASE_IDX - 1].date() - date.today()).days

        # 'save/update' row
        prev_row = row
        row_num += 1

    # save
    # configures new path for the exported file (append '_processed')
    new_path = dir_path + "/" + os.path.splitext(os.path.basename(file_path))[0] + "_processed.xlsx"
    # check to see if file is open in excel
    try:
        workbook.save(filename=new_path)
        response_label.configure(text="File Exported to: '" + new_path + "'", fg="green")
    except:
        response_label.configure(text="Failed to save file. Is the file currently open in Excel?", fg="red")
        return
    finally:
        # disable the buttons
        start_btn['state'] = tk.DISABLED


if __name__ == "__main__":
    main()
